"""
Professional Excel Export dengan Analitik
Format laporan perusahaan tingkat tinggi
"""
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, timedelta
from typing import Optional
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_db
from models import Product, Transaction, User
from auth import get_current_admin

router = APIRouter(prefix="/api/export/excel", tags=["excel-export"])


def create_professional_header(ws, title: str, subtitle: str = ""):
    """Create professional header for Excel report"""
    # Company Header
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "🏪 SISTEM KASIR KYUZU"
    cell.font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Report Title
    ws.merge_cells('A2:H2')
    cell = ws['A2']
    cell.value = title
    cell.font = Font(name='Arial', size=14, bold=True)
    cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 25
    
    # Subtitle/Date
    ws.merge_cells('A3:H3')
    cell = ws['A3']
    cell.value = subtitle or f"Dicetak: {datetime.now().strftime('%d %B %Y, %H:%M')}"
    cell.font = Font(name='Arial', size=10, italic=True)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.row_dimensions[3].height = 20


def apply_table_style(ws, start_row: int, end_row: int, num_cols: int, header_row: int = None):
    """Apply professional table styling"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header styling
    if header_row:
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # Data styling with alternating colors
    for row in range(start_row, end_row + 1):
        fill_color = "F2F2F2" if row % 2 == 0 else "FFFFFF"
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.font = Font(name='Arial', size=10)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')


def auto_adjust_columns(ws, min_width=10, max_width=50):
    """Auto-adjust column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


@router.get("/transactions")
async def export_transactions_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_admin)
):
    """Export transaksi ke Excel dengan analitik lengkap"""
    
    # Query data
    query = db.query(Transaction)
    if start_date:
        query = query.filter(Transaction.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.filter(Transaction.created_at <= datetime.fromisoformat(end_date))
    
    transactions = query.order_by(Transaction.created_at.desc()).all()
    
    # Create workbook
    wb = Workbook()
    
    # ===== SHEET 1: SUMMARY & ANALYTICS =====
    ws_summary = wb.active
    ws_summary.title = "📊 Ringkasan"
    
    create_professional_header(
        ws_summary,
        "LAPORAN KEUANGAN - RINGKASAN EKSEKUTIF",
        f"Periode: {start_date or 'Awal'} s/d {end_date or 'Sekarang'}"
    )
    
    # Calculate analytics
    total_transactions = len(transactions)
    total_revenue = sum(t.total for t in transactions)
    total_cost = sum(t.cost_total for t in transactions)
    total_profit = total_revenue - total_cost
    total_discount = sum(t.discount_amount for t in transactions)
    profit_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
    
    # Payment method breakdown
    payment_stats = {}
    for t in transactions:
        method = t.payment_method
        if method not in payment_stats:
            payment_stats[method] = {'count': 0, 'total': 0, 'profit': 0}
        payment_stats[method]['count'] += 1
        payment_stats[method]['total'] += t.total
        payment_stats[method]['profit'] += (t.total - t.cost_total)
    
    # Write summary data (starting row 5)
    row = 5
    ws_summary[f'A{row}'] = "METRIK UTAMA"
    ws_summary[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_summary[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws_summary.merge_cells(f'A{row}:B{row}')
    row += 1
    
    # Key metrics with color coding
    metrics = [
        ("Total Transaksi", total_transactions, "4472C4"),
        ("Total Pendapatan", f"Rp {total_revenue:,.0f}", "70AD47"),
        ("Total Modal", f"Rp {total_cost:,.0f}", "FFC000"),
        ("Laba Bersih", f"Rp {total_profit:,.0f}", "00B050" if total_profit > 0 else "C00000"),
        ("Margin Laba", f"{profit_margin:.1f}%", "00B050" if profit_margin > 10 else "FFC000"),
        ("Total Diskon", f"Rp {total_discount:,.0f}", "C00000"),
    ]
    
    for label, value, color in metrics:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'].font = Font(bold=True, color="FFFFFF")
        ws_summary[f'B{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws_summary[f'B{row}'].alignment = Alignment(horizontal='right')
        row += 1
    
    row += 2
    
    # Payment method breakdown table
    ws_summary[f'A{row}'] = "METODE PEMBAYARAN"
    ws_summary[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_summary[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws_summary.merge_cells(f'A{row}:D{row}')
    row += 1
    
    # Header for payment table
    headers = ["Metode", "Jumlah", "Total", "%"]
    for col, header in enumerate(headers, start=1):
        cell = ws_summary.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    start_payment_data = row
    for method, stats in payment_stats.items():
        percentage = (stats['total'] / total_revenue * 100) if total_revenue > 0 else 0
        ws_summary[f'A{row}'] = method.upper()
        ws_summary[f'B{row}'] = stats['count']
        ws_summary[f'C{row}'] = f"Rp {stats['total']:,.0f}"
        ws_summary[f'D{row}'] = f"{percentage:.1f}%"
        row += 1
    
    apply_table_style(ws_summary, start_payment_data, row-1, 4, start_payment_data-1)
    auto_adjust_columns(ws_summary)
    
    # ===== SHEET 2: DETAIL TRANSAKSI =====
    ws_detail = wb.create_sheet("📝 Detail Transaksi")
    
    create_professional_header(
        ws_detail,
        "DETAIL TRANSAKSI",
        f"Total: {total_transactions} transaksi"
    )
    
    # Headers (row 5)
    detail_headers = [
        "ID", "Tanggal", "Jam", "Kasir", "Pelanggan", 
        "Subtotal", "Diskon", "Total", "Modal", "Laba", "Margin %", "Metode"
    ]
    
    for col, header in enumerate(detail_headers, start=1):
        ws_detail.cell(row=5, column=col, value=header)
    
    # Data
    row = 6
    for t in transactions:
        profit = t.total - t.cost_total
        margin = (profit / t.total * 100) if t.total > 0 else 0
        
        ws_detail[f'A{row}'] = t.id
        ws_detail[f'B{row}'] = t.created_at.strftime("%Y-%m-%d")
        ws_detail[f'C{row}'] = t.created_at.strftime("%H:%M")
        ws_detail[f'D{row}'] = t.user.full_name if t.user else "-"
        ws_detail[f'E{row}'] = t.customer.name if t.customer else "-"
        ws_detail[f'F{row}'] = t.subtotal
        ws_detail[f'G{row}'] = t.discount_amount
        ws_detail[f'H{row}'] = t.total
        ws_detail[f'I{row}'] = t.cost_total
        ws_detail[f'J{row}'] = profit
        ws_detail[f'K{row}'] = margin
        ws_detail[f'L{row}'] = t.payment_method.upper()
        row += 1
    
    apply_table_style(ws_detail, 6, row-1, 12, 5)
    
    # Format currency columns
    for row_num in range(6, row):
        for col in ['F', 'G', 'H', 'I', 'J']:
            ws_detail[f'{col}{row_num}'].number_format = '#,##0'
        ws_detail[f'K{row_num}'].number_format = '0.0"%"'
    
    auto_adjust_columns(ws_detail)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"laporan_transaksi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/products")
async def export_products_excel(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_admin)
):
    """Export produk ke Excel dengan analitik inventori"""
    
    products = db.query(Product).filter(Product.is_active == True).order_by(Product.category, Product.name).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "📦 Inventori Produk"
    
    create_professional_header(
        ws,
        "LAPORAN INVENTORI PRODUK",
        f"Total: {len(products)} produk aktif"
    )
    
    # Summary stats
    total_value = sum(p.stock * p.price for p in products)
    total_cost_value = sum(p.stock * p.cost_price for p in products)
    low_stock_count = sum(1 for p in products if p.stock <= p.min_stock)
    
    ws['A5'] = "Total Nilai Stok:"
    ws['B5'] = f"Rp {total_value:,.0f}"
    ws['A6'] = "Total Nilai Modal:"
    ws['B6'] = f"Rp {total_cost_value:,.0f}"
    ws['A7'] = "Produk Stok Rendah:"
    ws['B7'] = low_stock_count
    
    for row in [5, 6, 7]:
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'B{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # Headers (row 9)
    headers = [
        "ID", "Barcode", "Nama", "Kategori", "Harga Jual", "Harga Modal", 
        "Margin", "Stok", "Min Stok", "Status", "Nilai Stok"
    ]
    
    for col, header in enumerate(headers, start=1):
        ws.cell(row=9, column=col, value=header)
    
    # Data
    row = 10
    for p in products:
        margin = ((p.price - p.cost_price) / p.price * 100) if p.price > 0 else 0
        status = "⚠️ RENDAH" if p.stock <= p.min_stock else "✅ OK"
        stock_value = p.stock * p.price
        
        ws[f'A{row}'] = p.id
        ws[f'B{row}'] = p.barcode or "-"
        ws[f'C{row}'] = p.name
        ws[f'D{row}'] = p.category
        ws[f'E{row}'] = p.price
        ws[f'F{row}'] = p.cost_price
        ws[f'G{row}'] = margin
        ws[f'H{row}'] = p.stock
        ws[f'I{row}'] = p.min_stock
        ws[f'J{row}'] = status
        ws[f'K{row}'] = stock_value
        
        # Color code status
        if p.stock <= p.min_stock:
            ws[f'J{row}'].font = Font(color="C00000", bold=True)
            ws[f'H{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        row += 1
    
    apply_table_style(ws, 10, row-1, 11, 9)
    
    # Format numbers
    for row_num in range(10, row):
        for col in ['E', 'F', 'K']:
            ws[f'{col}{row_num}'].number_format = '#,##0'
        ws[f'G{row_num}'].number_format = '0.0"%"'
    
    auto_adjust_columns(ws)
    
    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"laporan_produk_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
