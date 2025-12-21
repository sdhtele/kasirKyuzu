"""
Professional Excel Export with Charts and Formatting
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, PieChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict
import io


def create_professional_transaction_report(transactions: List, summary_data: Dict) -> bytes:
    """Create professional transaction report with charts and formatting"""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # ========== SUMMARY SHEET ==========
    ws_summary = wb.create_sheet("Dashboard Summary", 0)
    
    # Title
    ws_summary['A1'] = "LAPORAN KEUANGAN"
    ws_summary['A1'].font = Font(size=24, bold=True, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    ws_summary['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.merge_cells('A1:F1')
    ws_summary.row_dimensions[1].height = 40
    
    # Date range
    ws_summary['A2'] = f"Periode: {datetime.now().strftime('%d %B %Y')}"
    ws_summary['A2'].font = Font(size=12, italic=True)
    ws_summary.merge_cells('A2:F2')
    
    # Key Metrics Header
    ws_summary['A4'] = "KEY PERFORMANCE INDICATORS"
    ws_summary['A4'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_summary['A4'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    ws_summary.merge_cells('A4:F4')
    ws_summary['A4'].alignment = Alignment(horizontal="center")
    
    # KPIs
    kpi_row = 6
    kpis = [
        ("Total Pendapatan", summary_data.get('total_revenue', 0), "4F46E5"),
        ("Total Transaksi", summary_data.get('total_transactions', 0), "10B981"),
        ("Rata-rata/Transaksi", summary_data.get('avg_transaction', 0), "F59E0B"),
        ("Total Laba", summary_data.get('total_profit', 0), "EF4444"),
    ]
    
    for idx, (label, value, color) in enumerate(kpis, start=1):
        cell_label = ws_summary.cell(row=kpi_row, column=idx*2-1)
        cell_value = ws_summary.cell(row=kpi_row, column=idx*2)
        
        cell_label.value = label
        cell_label.font = Font(bold=True, size=11)
        cell_label.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        
        cell_value.value = value if isinstance(value, int) else f"Rp {value:,.0f}"
        cell_value.font = Font(size=14, bold=True, color=color)
        cell_value.alignment = Alignment(horizontal="right")
    
    # ========== TRANSACTIONS SHEET ==========
    ws_trans = wb.create_sheet("Detail Transaksi", 1)
    
    # Headers
    headers = ["ID", "Tanggal", "Waktu", "Kasir", "Pelanggan", "Subtotal", "Diskon", 
               "Total", "Modal", "Laba", "Margin %", "Metode Bayar"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws_trans.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Data rows
    for row_idx, t in enumerate(transactions, start=2):
        values = [
            t.id,
            t.created_at.strftime("%d/%m/%Y"),
            t.created_at.strftime("%H:%M"),
            t.user.full_name if t.user else "-",
            t.customer.name if t.customer else "Umum",
            t.subtotal,
            t.discount_amount,
            t.total,
            t.cost_total,
            t.total - t.cost_total,
            round((t.total - t.cost_total) / t.total * 100, 1) if t.total > 0 else 0,
            t.payment_method.upper()
        ]
        
        for col, value in enumerate(values, start=1):
            cell = ws_trans.cell(row=row_idx, column=col)
            cell.value = value
            
            # Styling
            if col >= 6 and col <= 10:  # Money columns
                cell.number_format = '#,##0'
            if col == 11:  # Percentage
                cell.number_format = '0.0"%"'
            
            # Alternating rows
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            
            cell.border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws_trans.column_dimensions[get_column_letter(col)].width = 15
    
    # Freeze panes
    ws_trans.freeze_panes = "A2"
    
    # ========== ANALYTICS SHEET ==========
    ws_analytics = wb.create_sheet("Analitik", 2)
    
    # Add charts section header
    ws_analytics['A1'] = "ANALISIS PERFORMA PENJUALAN"
    ws_analytics['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_analytics['A1'].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    ws_analytics.merge_cells('A1:H1')
    ws_analytics['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Payment method summary
    ws_analytics['A3'] = "Metode Pembayaran"
    ws_analytics['A3'].font = Font(bold=True, size=12)
    ws_analytics['B3'] = "Jumlah Transaksi"
    ws_analytics['B3'].font = Font(bold=True, size=12)
    ws_analytics['C3'] = "Total"
    ws_analytics['C3'].font = Font(bold=True, size=12)
    
    payment_data = {}
    for t in transactions:
        method = t.payment_method.upper()
        if method not in payment_data:
            payment_data[method] = {'count': 0, 'total': 0}
        payment_data[method]['count'] += 1
        payment_data[method]['total'] += t.total
    
    row = 4
    for method, data in payment_data.items():
        ws_analytics.cell(row=row, column=1).value = method
        ws_analytics.cell(row=row, column=2).value = data['count']
        ws_analytics.cell(row=row, column=3).value = data['total']
        ws_analytics.cell(row=row, column=3).number_format = '#,##0'
        row += 1
    
    # Create pie chart for payment methods
    if len(payment_data) > 0:
        pie_chart = PieChart()
        pie_chart.title = "Distribusi Metode Pembayaran"
        pie_chart.style = 10
        
        labels = Reference(ws_analytics, min_col=1, min_row=4, max_row=3+len(payment_data))
        data_ref = Reference(ws_analytics, min_col=3, min_row=3, max_row=3+len(payment_data))
        
        pie_chart.add_data(data_ref, titles_from_data=True)
        pie_chart.set_categories(labels)
        pie_chart.height = 10
        pie_chart.width = 15
        
        ws_analytics.add_chart(pie_chart, "E3")
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_professional_product_report(products: List, summary_data: Dict) -> bytes:
    """Create professional product/inventory report"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # ========== SUMMARY SHEET ==========
    ws_summary = wb.create_sheet("Dashboard Inventori", 0)
    
    # Title
    ws_summary['A1'] = "LAPORAN INVENTORI"
    ws_summary['A1'].font = Font(size=24, bold=True, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    ws_summary['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.merge_cells('A1:F1')
    ws_summary.row_dimensions[1].height = 40
    
    # KPIs
    kpi_row = 4
    kpis = [
        ("Total Produk", len(products), "4F46E5"),
        ("Total Stok", sum(p.stock for p in products), "10B981"),
        ("Stok Rendah", len([p for p in products if p.stock <= p.min_stock]), "F59E0B"),
        ("Nilai Inventori", sum(p.stock * p.price for p in products), "EF4444"),
    ]
    
    for idx, (label, value, color) in enumerate(kpis, start=1):
        cell_label = ws_summary.cell(row=kpi_row, column=idx*2-1)
        cell_value = ws_summary.cell(row=kpi_row+1, column=idx*2-1)
        
        cell_label.value = label
        cell_label.font = Font(bold=True, size=11)
        cell_label.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        
        cell_value.value = value if idx != 4 else f"Rp {value:,.0f}"
        cell_value.font = Font(size=16, bold=True, color=color)
        cell_value.alignment = Alignment(horizontal="center")
    
    # ========== PRODUCTS SHEET ==========
    ws_products = wb.create_sheet("Detail Produk", 1)
    
    headers = ["ID", "Nama Produk", "Kategori", "Harga Jual", "Harga Modal", 
               "Stok", "Min Stok", "Status", "Nilai Stok", "Profit Margin"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws_products.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, p in enumerate(products, start=2):
        stock_value = p.stock * p.price
        profit_margin = ((p.price - p.cost_price) / p.price * 100) if p.price > 0 else 0
        status = "⚠️ RENDAH" if p.stock <= p.min_stock else "✅ OK"
        
        values = [
            p.id,
            p.name,
            p.category,
            p.price,
            p.cost_price,
            p.stock,
            p.min_stock,
            status,
            stock_value,
            round(profit_margin, 1)
        ]
        
        for col, value in enumerate(values, start=1):
            cell = ws_products.cell(row=row_idx, column=col)
            cell.value = value
            
            if col in [4, 5, 9]:  # Money columns
                cell.number_format = '#,##0'
            if col == 10:  # Percentage
                cell.number_format = '0.0"%"'
            
            # Color code low stock
            if col == 8 and "RENDAH" in str(value):
                cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                cell.font = Font(color="92400E", bold=True)
            
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    # Auto-width
    for col in range(1, len(headers) + 1):
        ws_products.column_dimensions[get_column_letter(col)].width = 15
    
    ws_products.freeze_panes = "A2"
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
